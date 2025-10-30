"""
Real-time Excel Writer for NCIIPC PS-02 Submission

Provides thread-safe Excel file operations for continuous submission data collection.

Stage 1 (Crawler): Writes 18 columns from crawler/ChromaDB metadata
Stage 2 (AIML): Updates columns 6 & 20 with ML detection verdicts

Thread-safe with file locking to handle concurrent writes.
"""

import os
import json
import threading
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import logging
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)


class RealtimeExcelWriter:
    """Thread-safe Excel writer for real-time submission data"""

    # Annexure B: Required 20 columns
    REQUIRED_COLUMNS = [
        'Application_ID',
        'Source of detection',
        'Identified Phishing/Suspected Domain Name',
        'Corresponding CSE Domain Name',
        'Critical Sector Entity Name',
        'Phishing/Suspected Domains',  # Column 6 - Updated by AIML
        'Domain Registration Date',
        'Registrar Name',
        'Registrant Name or Registrant Organisation',
        'Registrant Country',
        'Name Servers',
        'Hosting IP',
        'Hosting ISP',
        'Hosting Country',
        'DNS Records (if any)',
        'Evidence file name',
        'Date of detection (DD-MM-YYYY)',
        'Time of detection (HH-MM-SS)',
        'Date of Post (If detection is from Source: social media)',
        'Remarks (If any)'  # Column 20 - Updated by AIML
    ]

    def __init__(self, excel_path: str, application_id: str = None):
        """
        Initialize Excel writer

        Args:
            excel_path: Path to Excel file
            application_id: Application ID (from env or parameter)
        """
        self.excel_path = Path(excel_path)
        self.application_id = application_id or os.getenv('SUBMISSION_APP_ID', 'UNKNOWN')
        self.lock = threading.Lock()

        # CSE mapping (domain -> organization)
        self.cse_mapping = self._load_cse_mapping()

        # Initialize Excel file if doesn't exist
        if not self.excel_path.exists():
            self._initialize_excel()
            logger.info(f"Initialized Excel file: {self.excel_path}")
        else:
            logger.info(f"Using existing Excel file: {self.excel_path}")

    def _load_cse_mapping(self) -> Dict[str, str]:
        """Load CSE domain to organization name mapping"""
        cse_map = {}

        # Try loading from baseline profile
        baseline_path = Path('data/training/cse_baseline_profile.json')
        if baseline_path.exists():
            try:
                with open(baseline_path, 'r') as f:
                    baseline = json.load(f)
                    if 'cse_organizations' in baseline:
                        cse_map = baseline['cse_organizations']
                    elif 'domains' in baseline:
                        for domain in baseline['domains']:
                            org_name = self._extract_org_name(domain)
                            cse_map[domain] = org_name
            except Exception as e:
                logger.warning(f"Could not load CSE mapping: {e}")

        # Default mappings
        default_mappings = {
            'onlinesbi.com': 'State Bank of India',
            'sbi.co.in': 'State Bank of India',
            'icicibank.com': 'ICICI Bank',
            'hdfcbank.com': 'HDFC Bank',
            'axisbank.com': 'Axis Bank',
            'pnb.co.in': 'Punjab National Bank',
            'bank.in': 'Banking Sector',
            'airtel.in': 'Airtel',
            'airtel.com': 'Bharti Airtel Limited',
            'bsnl.co.in': 'BSNL',
            'uidai.gov.in': 'UIDAI',
            'epfindia.gov.in': 'EPFO',
            'irctc.co.in': 'IRCTC'
        }

        for domain, org in default_mappings.items():
            if domain not in cse_map:
                cse_map[domain] = org

        return cse_map

    def _extract_org_name(self, domain: str) -> str:
        """Extract organization name from domain"""
        base = domain.split('.')[0]
        return base.upper()

    def _initialize_excel(self):
        """Create new Excel file with headers"""
        try:
            # Create parent directory if needed
            self.excel_path.parent.mkdir(parents=True, exist_ok=True)

            # Create workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "Submission"

            # Write headers (bold font)
            bold_font = Font(bold=True)
            for col_idx, column_name in enumerate(self.REQUIRED_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=column_name)
                cell.font = bold_font

            # Auto-adjust column widths
            for col_idx in range(1, len(self.REQUIRED_COLUMNS) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

            wb.save(self.excel_path)
            logger.info(f"Created Excel file with headers: {self.excel_path}")

        except Exception as e:
            logger.error(f"Failed to initialize Excel file: {e}")
            raise

    def get_cse_domain(self, metadata: Dict) -> str:
        """Extract corresponding CSE domain from metadata"""
        cse_domain = metadata.get('seed_registrable', '')
        if not cse_domain:
            cse_domain = metadata.get('original_domain', '')
        if not cse_domain:
            cse_id = metadata.get('cse_id', '')
            if cse_id and cse_id not in ['Unknown', 'BULK_IMPORT', 'URL from user']:
                cse_domain = cse_id
        return cse_domain or 'N/A'

    def get_cse_organization(self, cse_domain: str) -> str:
        """Get organization name for CSE domain"""
        if cse_domain == 'N/A' or not cse_domain:
            return 'N/A'

        # Direct match
        if cse_domain in self.cse_mapping:
            return self.cse_mapping[cse_domain]

        # Partial match
        for known_domain, org in self.cse_mapping.items():
            if known_domain in cse_domain or cse_domain in known_domain:
                return org

        # Extract from domain name
        return self._extract_org_name(cse_domain)

    def format_dns_records(self, metadata: Dict) -> str:
        """Format DNS records from metadata"""
        dns_records = []

        # Parse DNS JSON if available
        dns_str = metadata.get('dns', '')
        if dns_str:
            try:
                dns_data = json.loads(dns_str) if isinstance(dns_str, str) else dns_str
                for record_type in ['A', 'AAAA', 'MX', 'NS', 'CNAME']:
                    if record_type in dns_data and dns_data[record_type]:
                        records = dns_data[record_type]
                        if isinstance(records, list) and records:
                            dns_records.append(f"{record_type}: {', '.join(map(str, records[:3]))}")
            except:
                pass

        # Fallback to individual fields
        if not dns_records:
            a_count = metadata.get('a_count', 0)
            mx_count = metadata.get('mx_count', 0)
            ns_count = metadata.get('ns_count', 0)
            if a_count or mx_count or ns_count:
                dns_records.append(f"A: {a_count}, MX: {mx_count}, NS: {ns_count}")

        return '; '.join(dns_records) if dns_records else 'N/A'

    def get_name_servers(self, metadata: Dict) -> str:
        """Extract name servers from DNS metadata"""
        dns_str = metadata.get('dns', '')
        if dns_str:
            try:
                dns_data = json.loads(dns_str) if isinstance(dns_str, str) else dns_str
                ns_list = dns_data.get('NS', [])
                if ns_list:
                    return ', '.join(ns_list[:3])  # Limit to 3
            except:
                pass

        ns_count = metadata.get('ns_count', 0)
        return f"{ns_count} nameserver(s)" if ns_count > 0 else 'N/A'

    def get_hosting_ip(self, metadata: Dict) -> str:
        """Extract hosting IP address"""
        # Try DNS A records first
        dns_str = metadata.get('dns', '')
        if dns_str:
            try:
                dns_data = json.loads(dns_str) if isinstance(dns_str, str) else dns_str
                a_records = dns_data.get('A', [])
                if a_records:
                    return ', '.join(map(str, a_records[:2]))  # First 2 IPs
            except:
                pass

        return 'N/A'

    def parse_timestamp(self, timestamp_str: str) -> tuple:
        """Parse ISO timestamp to (DD-MM-YYYY, HH-MM-SS)"""
        if not timestamp_str:
            return ('N/A', 'N/A')

        try:
            # Handle various timestamp formats
            if 'T' in timestamp_str:
                dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            else:
                dt = datetime.fromisoformat(timestamp_str)

            date_str = dt.strftime('%d-%m-%Y')
            time_str = dt.strftime('%H-%M-%S')
            return (date_str, time_str)
        except Exception as e:
            logger.warning(f"Failed to parse timestamp {timestamp_str}: {e}")
            return ('N/A', 'N/A')

    def calculate_registration_date(self, metadata: Dict) -> str:
        """Calculate domain registration date from age"""
        domain_age_days = metadata.get('domain_age_days', 0)

        if domain_age_days and domain_age_days > 0:
            try:
                # Calculate registration date from age
                reg_date = datetime.now() - pd.Timedelta(days=int(domain_age_days))
                return reg_date.strftime('%d-%m-%Y')
            except:
                pass

        # Try first_seen timestamp
        first_seen = metadata.get('first_seen', '')
        if first_seen:
            date_part, _ = self.parse_timestamp(first_seen)
            if date_part != 'N/A':
                return date_part

        return 'N/A'

    def build_crawler_row(self, metadata: Dict) -> Dict:
        """
        Build Excel row from crawler metadata (Stage 1)

        Columns 1-19 (except column 6 which AIML updates later)
        Column 6 is set to "PENDING_AIML"
        Column 20 is set to "Awaiting AIML analysis"
        """
        domain = metadata.get('registrable') or metadata.get('domain', 'N/A')

        # Parse timestamp
        first_seen = metadata.get('first_seen', '')
        detection_date, detection_time = self.parse_timestamp(first_seen)

        # Get CSE info
        cse_domain = self.get_cse_domain(metadata)
        cse_org = self.get_cse_organization(cse_domain)

        # Evidence filename (placeholder for now, updated when screenshot exists)
        evidence_filename = f"{cse_org.replace(' ', '')}_{domain.replace('.', '_')}.pdf"

        # Build row with all 20 columns
        row = {
            'Application_ID': self.application_id,
            'Source of detection': 'crawler_pipeline',
            'Identified Phishing/Suspected Domain Name': domain,
            'Corresponding CSE Domain Name': cse_domain,
            'Critical Sector Entity Name': cse_org,
            'Phishing/Suspected Domains': 'PENDING_AIML',  # Column 6 - Updated by AIML later
            'Domain Registration Date': self.calculate_registration_date(metadata),
            'Registrar Name': metadata.get('registrar', 'N/A'),
            'Registrant Name or Registrant Organisation': 'N/A',  # Privacy protected
            'Registrant Country': metadata.get('country', 'N/A'),
            'Name Servers': self.get_name_servers(metadata),
            'Hosting IP': self.get_hosting_ip(metadata),
            'Hosting ISP': metadata.get('asn_org', 'N/A'),
            'Hosting Country': metadata.get('country', 'N/A'),
            'DNS Records (if any)': self.format_dns_records(metadata),
            'Evidence file name': evidence_filename,
            'Date of detection (DD-MM-YYYY)': detection_date,
            'Time of detection (HH-MM-SS)': detection_time,
            'Date of Post (If detection is from Source: social media)': 'N/A',
            'Remarks (If any)': 'Awaiting AIML analysis'  # Column 20 - Updated by AIML later
        }

        return row

    def append_batch_to_excel(self, metadata_batch: List[Dict]):
        """
        Append batch of crawler data to Excel (Stage 1)

        Thread-safe batch append operation.
        """
        if not metadata_batch:
            return

        with self.lock:
            try:
                # Build rows
                rows = []
                for metadata in metadata_batch:
                    try:
                        row = self.build_crawler_row(metadata)
                        rows.append(row)
                    except Exception as e:
                        logger.warning(f"Failed to build row for domain {metadata.get('registrable', 'unknown')}: {e}")
                        continue

                if not rows:
                    return

                # Load existing Excel
                wb = load_workbook(self.excel_path)
                ws = wb.active

                # Find next empty row
                next_row = ws.max_row + 1

                # Append rows
                for row_data in rows:
                    for col_idx, column_name in enumerate(self.REQUIRED_COLUMNS, 1):
                        value = row_data.get(column_name, 'N/A')
                        ws.cell(row=next_row, column=col_idx, value=value)
                    next_row += 1

                # Save
                wb.save(self.excel_path)
                logger.info(f"Appended {len(rows)} rows to Excel (Stage 1 - Crawler data)")

            except Exception as e:
                logger.error(f"Failed to append batch to Excel: {e}", exc_info=True)

    def update_aiml_verdict(self, domain: str, verdict: str, confidence: float, reason: str):
        """
        Update Excel with AIML verdict (Stage 2)

        Updates columns 6 and 20 for the specified domain.
        Thread-safe update operation.
        """
        with self.lock:
            try:
                # Load Excel
                wb = load_workbook(self.excel_path)
                ws = wb.active

                # Find row by domain name (column 3)
                domain_col_idx = 3  # "Identified Phishing/Suspected Domain Name"
                verdict_col_idx = 6  # "Phishing/Suspected Domains"
                remarks_col_idx = 20  # "Remarks (If any)"

                row_found = False
                for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                    cell_value = ws.cell(row=row_idx, column=domain_col_idx).value
                    if cell_value == domain:
                        # Update verdict column (6)
                        ws.cell(row=row_idx, column=verdict_col_idx, value=verdict.upper())

                        # Update remarks column (20)
                        remarks = f"Confidence: {confidence:.2f}, {reason}"[:200]
                        ws.cell(row=row_idx, column=remarks_col_idx, value=remarks)

                        row_found = True
                        break

                if row_found:
                    wb.save(self.excel_path)
                    logger.info(f"Updated Excel verdict for domain: {domain} -> {verdict}")
                else:
                    logger.warning(f"Domain not found in Excel for verdict update: {domain}")

            except Exception as e:
                logger.error(f"Failed to update Excel verdict for {domain}: {e}", exc_info=True)


# Global instance (lazy initialization)
_excel_writer = None
_writer_lock = threading.Lock()


def get_excel_writer() -> Optional[RealtimeExcelWriter]:
    """
    Get or create global Excel writer instance

    Returns:
        RealtimeExcelWriter instance or None if disabled
    """
    global _excel_writer

    # Check if feature is enabled
    enabled = os.getenv('EXCEL_REALTIME_ENABLED', 'false').lower() == 'true'
    if not enabled:
        return None

    # Lazy initialization
    if _excel_writer is None:
        with _writer_lock:
            if _excel_writer is None:  # Double-check
                excel_path = os.getenv('EXCEL_OUTPUT_PATH', '/out/PS-02_Live_Submission.xlsx')
                app_id = os.getenv('SUBMISSION_APP_ID', 'UNKNOWN')

                try:
                    _excel_writer = RealtimeExcelWriter(excel_path, app_id)
                    logger.info(f"Initialized global Excel writer: {excel_path}")
                except Exception as e:
                    logger.error(f"Failed to initialize Excel writer: {e}")
                    return None

    return _excel_writer
